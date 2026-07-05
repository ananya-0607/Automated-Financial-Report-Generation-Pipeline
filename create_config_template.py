"""
create_config_template.py
Creates a blank slides_config.xlsx with dropdowns, docs, and color palette.
Only needs the PowerPoint template as input.

Usage:
  python create_config_template.py
  python create_config_template.py --template input/other_template.pptx
"""

import argparse
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from pptx import Presentation


# ── Master column list — identical order to COLS in all runner scripts ──
# CPI / Light House / AMBIT runners use the same list; this file must stay in sync.
COL_DEFS = [
    # ── Identity ─────────────────────────────────────────────────────
    ("slide_number",          "Slide Number",             "Which slide this chart belongs to (1, 2, 3...)"),
    ("slide_title",           "Slide Title",              "Title shown on the slide (e.g., 'Banking Indicators')"),
    ("layout_name",           "Template Layout",          "PowerPoint layout name — pick from dropdown"),
    # ── Primary widget ───────────────────────────────────────────────
    ("widget_id",             "Primary Widget ID",        "Thurro widget ID for the main data series"),
    ("widget_id_older",       "Primary Widget (Older)",   "Older widget ID to stitch longer history (0 = none)"),
    ("Fetching_Method",       "Fetching Method",          "API = Thurro API; ClickHouse = provide SQL query"),
    # ── Primary axis ─────────────────────────────────────────────────
    ("y1_label",              "Primary Y-Axis Label",     "Label for the primary axis (shown in legend)"),
    ("y1_chart_type",         "Primary Chart Type",       "bar | line | stacked_bar"),
    ("y1_divisor",            "Primary Divisor",          "Divide raw values by this (1 = none, 1e9 = billions)"),
    ("bar_color",             "Primary Color",            "Hex color for primary series — pick from Color Palette sheet"),
    ("y_min",                 "Y-Axis Min",               "Minimum value for primary y-axis"),
    ("y_max",                 "Y-Axis Max",               "Maximum value for primary y-axis"),
    ("y_breaks",              "Y-Axis Breaks",            "Number of gridline intervals on primary y-axis"),
    # ── Secondary axis ───────────────────────────────────────────────
    ("y2_widget_id",          "Secondary Widget ID",      "Widget ID for secondary axis line (0 = no secondary axis)"),
    ("y2_widget_id_older",    "Secondary Widget (Older)", "Older widget for secondary axis stitching (0 = none)"),
    ("y2_label",              "Secondary Y-Axis Label",   "Label for secondary axis line"),
    ("y2_chart_type",         "Secondary Chart Type",     "Chart type for secondary axis (always line)"),
    ("y2_divisor",            "Secondary Divisor",        "Divisor for secondary axis values"),
    ("y2_color",              "Secondary Color",          "Hex color for secondary line — pick from dropdown"),
    ("y2_min",                "Secondary Y-Axis Min",     "Minimum for secondary y-axis (0 if unused)"),
    ("y2_max",                "Secondary Y-Axis Max",     "Maximum for secondary y-axis (0 if unused)"),
    ("y2_breaks",             "Secondary Y-Axis Breaks",  "Number of gridline intervals on secondary y-axis"),
    ("y2_start_date",         "Secondary Start Date",     "Optional start date for Y2 data (YYYY-MM-DD, blank = all)"),
    # ── X-axis / date filtering ──────────────────────────────────────
    ("x_grouping",            "X-Axis Grouping",          "MY = monthly, FY = fiscal year, QY = quarterly"),
    ("x_interval",            "X-Axis Label Interval",    "Show label every N periods (e.g., 12 = yearly for monthly data)"),
    ("last_n_months",         "Last N Months",            "Show only last N months (0 = show all available data)"),
    ("start_year",            "Start FY Year",            "Filter: fiscal year start (0 = no filter)"),
    ("end_year",              "End FY Year",              "Filter: fiscal year end (0 = no filter)"),
    # ── Calculations ─────────────────────────────────────────────────
    ("growth_type",           "Growth Calculation",       "none | yoy | mom | qoq"),
    ("aggregate",             "Aggregation",              "none | ytd_fy | eo_q | eo_fy"),
    # ── Display toggles ──────────────────────────────────────────────
    ("show_values_all",        "Show All Values",          "TRUE = show label on EVERY data point across all chart types"),
    ("show_values_latest",    "Show Latest Value",        "TRUE = show label on LATEST data point only (overrides show_values_all if set)"),
    ("show_grid",             "Show Grid Lines",          "TRUE = display grid lines"),
    # ── Source / notes ───────────────────────────────────────────────
    ("source_override",       "Source Override",          "Custom source text (blank = auto-generated)"),
    ("notes",                 "Notes",                    "Internal notes — not shown on the slide"),
    # ── Matched titles (auto-filled by match step) ───────────────────
    ("matched_chart_title",   "Matched Chart Title",      "Chart title matched from AI or template"),
    ("matched_chart_source",  "Matched Chart Source",     "Source string matched from template"),
    # ── Chart text ───────────────────────────────────────────────────
    ("chart_note",            "Chart Note",               "Note text placed below Source (disclaimers, methodology)"),
    ("chart_content",         "Chart Content",            "Content text placed below Note"),
    ("subtitle_mode",         "Subtitle Date Mode",       "FY | CY | FY_DATE | MONTH_ONLY"),
    ("subtitle_prefix",       "Subtitle Prefix",          "Manual text before date range (e.g., 'Monthly trade (USD bn)')"),
    # ── Multi-series (stack / multi-line) ────────────────────────────
    ("stack_widgets",         "Stacked Widget IDs",       "Pipe-separated extra widget IDs: 123|456|789"),
    ("stack_labels",          "Stacked Labels",           "Pipe-separated labels (primary first): Label1|Label2|Label3"),
    ("stack_colors",          "Stacked Colors",           "Pipe-separated hex colors (primary first): #AAA|#BBB|#CCC"),
    ("stack_signs",           "Stacked Signs",            "Pipe-separated pos/neg per series: pos|neg|pos"),
    # ── Legend / axis overrides ──────────────────────────────────────
    ("legend_loc",            "Legend Position",          "top | bottom | right | none | blank = use theme default"),
    ("legend_ncol",           "Legend Columns",           "Number of columns in the legend (0 = auto)"),
    ("legend_nrow",           "Legend Rows",              "Number of rows in the legend (0 = auto)"),
    ("x_axis_reverse",        "Reverse X-Axis",           "TRUE = newest data on left; blank/FALSE = oldest→newest"),
    ("x_tick_angle",          "X-Axis Label Angle",       "Rotation of x-axis tick labels in degrees (0=horizontal, 45=angled, 90=vertical). Blank = use YAML default (-1 = auto)."),
    ("show_y2_axis",          "Show Secondary Y-Axis",    "TRUE = show right-side Y2 axis spine. FALSE = hide. Blank = use YAML default."),
    ("drop_last_period",      "Drop Last Period",          "TRUE = remove the most recent data point before plotting (useful when latest month is incomplete). Blank/FALSE = show all data."),
    # ── Matched slide text (auto-filled by match step) ───────────────
    ("matched_slide_heading",     "Matched Slide Heading",     "Slide heading matched from template"),
    ("matched_slide_sub_heading", "Matched Slide Sub-Heading", "Slide sub-heading matched from template"),
]

COLORS = {
    "#FFC125": "NIIF Gold",
    "#EE7600": "NIIF Orange",
    "#FFD39B": "Light Gold",
    "#FFA54F": "Medium Orange",
    "#E0E0E0": "Light Gray",
    "#888888": "Mid Gray",
    "#7A7A7A": "Dark Gray",
    "#EE9572": "Salmon",
    "#BF6E00": "Dark Gold",
}


def extract_layouts(template_path):
    prs = Presentation(template_path)
    names = []
    for master in prs.slide_masters:
        for l in master.slide_layouts:
            names.append(l.name)
    return sorted(set(names))


def create_template(template_path, output_path):
    layouts = extract_layouts(template_path)
    layout_list = ",".join(layouts)
    color_list = ",".join(COLORS.keys())

    wb = Workbook()

    # ═══════════════ Sheet 1: slides (blank) ═══════════════
    ws = wb.active
    ws.title = "slides"

    hdr_font = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="333333")
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Border(bottom=Side("thin", color="CCCCCC"),
                  right=Side("thin", color="CCCCCC"))

    for ci, (col_name, friendly, desc) in enumerate(COL_DEFS, 1):
        cell = ws.cell(row=1, column=ci, value=col_name)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = hdr_align
        cell.border = thin
        cell.comment = Comment(f"{friendly}\n\n{desc}", "Config Guide")

    widths = {
        "slide_number": 8, "slide_title": 22, "layout_name": 28,
        "widget_id": 14, "y1_label": 28, "y1_chart_type": 14,
        "bar_color": 12, "stack_widgets": 30, "stack_labels": 35,
        "stack_colors": 30, "chart_note": 25, "chart_content": 25,
        "notes": 18,
    }
    for ci, (col_name, _, _) in enumerate(COL_DEFS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = widths.get(col_name, 14)

    ws.freeze_panes = "A2"

    # Dropdowns
    max_row = 200
    dvs = {
        "y1_chart_type": DataValidation(type="list", formula1='"bar,line,stacked_bar"', allow_blank=True),
        "y2_chart_type": DataValidation(type="list", formula1='"line,"', allow_blank=True),
        "x_grouping":    DataValidation(type="list", formula1='"MY,FY"', allow_blank=True),
        "aggregate":     DataValidation(type="list", formula1='"none,ytd_fy"', allow_blank=True),
        "growth_type":   DataValidation(type="list", formula1='"none,yoy,mom,qoq"', allow_blank=True),
        "show_values_all":    DataValidation(type="list", formula1='"FALSE,TRUE"', allow_blank=True),
        "show_values_latest": DataValidation(type="list", formula1='"FALSE,TRUE"', allow_blank=True),
        "bar_color":     DataValidation(type="list", formula1=f'"{color_list}"', allow_blank=True),
        "y2_color":      DataValidation(type="list", formula1=f'"{color_list}"', allow_blank=True),
        "subtitle_mode": DataValidation(type="list", formula1='"FY,CY,FY_DATE,MONTH_ONLY"', allow_blank=True),
        "legend_loc":    DataValidation(type="list", formula1='"top,bottom,right,none"', allow_blank=True),
        "x_axis_reverse":DataValidation(type="list", formula1='"FALSE,TRUE"', allow_blank=True),
        "show_y2_axis":  DataValidation(type="list", formula1='"FALSE,TRUE"', allow_blank=True),
        "drop_last_period": DataValidation(type="list", formula1='"FALSE,TRUE"', allow_blank=True),
    }
    for dv in dvs.values():
        ws.add_data_validation(dv)
    for ci, (col_name, _, _) in enumerate(COL_DEFS, 1):
        if col_name in dvs:
            cl = get_column_letter(ci)
            dvs[col_name].add(f"{cl}2:{cl}{max_row}")

    # ═══════════════ Sheet 2: Column Guide ═══════════════
    ws2 = wb.create_sheet("Column Guide")
    for ci, hdr in enumerate(["Column Name", "Friendly Name", "Description"], 1):
        c = ws2.cell(row=1, column=ci, value=hdr)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="333333")
    for ri, (col_name, friendly, desc) in enumerate(COL_DEFS, 2):
        ws2.cell(row=ri, column=1, value=col_name).font = Font(name="Calibri", bold=True, size=10)
        ws2.cell(row=ri, column=2, value=friendly).font = Font(name="Calibri", size=10, color="EE7600")
        ws2.cell(row=ri, column=3, value=desc).font = Font(name="Calibri", size=10)
    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 26
    ws2.column_dimensions["C"].width = 70

    # ═══════════════ Sheet 3: Color Palette ═══════════════
    ws3 = wb.create_sheet("Color Palette")
    for ci, hdr in enumerate(["Hex Code", "Name", "Preview"], 1):
        c = ws3.cell(row=1, column=ci, value=hdr)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="333333")
    for ri, (hx, name) in enumerate(COLORS.items(), 2):
        ws3.cell(row=ri, column=1, value=hx)
        ws3.cell(row=ri, column=2, value=name)
        fc = hx.lstrip("#")
        c = ws3.cell(row=ri, column=3)
        c.fill = PatternFill("solid", fgColor=fc)
    ws3.column_dimensions["A"].width = 14
    ws3.column_dimensions["B"].width = 18
    ws3.column_dimensions["C"].width = 14

    # ═══════════════ Sheet 4: Layouts ═══════════════
    ws4 = wb.create_sheet("Layouts")
    for ci, hdr in enumerate(["Layout Name", "Charts Supported"], 1):
        c = ws4.cell(row=1, column=ci, value=hdr)
        c.font = Font(bold=True, size=11, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="333333")
    hints = {"1 ": "1 chart", "2 ": "2 charts", "3 ": "3 charts",
             "4 ": "4 charts", "BIGGER": "1 chart (large)",
             "SMALLER": "1 chart + notes"}
    for ri, name in enumerate(layouts, 2):
        ws4.cell(row=ri, column=1, value=name)
        hint = ""
        for k, v in hints.items():
            if k in name.upper():
                hint = v
                break
        ws4.cell(row=ri, column=2, value=hint)
    ws4.column_dimensions["A"].width = 35
    ws4.column_dimensions["B"].width = 20

    # layout_name dropdown — use range ref (no 255-char limit)
    n_layouts = len(layouts)
    layout_dv = DataValidation(
        type="list",
        formula1=f"=Layouts!$A$2:$A${n_layouts + 1}",
        allow_blank=True,
    )
    ws.add_data_validation(layout_dv)
    for ci, (col_name, _, _) in enumerate(COL_DEFS, 1):
        if col_name == "layout_name":
            cl = get_column_letter(ci)
            layout_dv.add(f"{cl}2:{cl}{max_row}")
            break

    wb.save(output_path)
    print(f"\n{'='*50}")
    print(f"Created: {output_path}")
    print(f"  {len(layouts)} layouts from {template_path}")
    print(f"  {len(COL_DEFS)} columns with dropdowns & docs")
    print(f"  {len(COLORS)} colors in palette")
    print(f"  Ready to fill — just add rows!")
    print(f"{'='*50}")


def load_project_paths(config_path, project):
    """Read the YAML and return the paths block for one project."""
    import yaml
    with open(config_path, encoding="utf-8") as f:
        cfg = yaml.safe_load(f)
    projects = cfg["default"]["projects"]
    if project not in projects:
        raise KeyError(f"Project '{project}' not found. Available: {list(projects)}")
    return projects[project]


if __name__ == "__main__":
    CONFIG_PATH="D:\Santonu\Desktop\ADQvest\Error files\Modified(corr)\R_PPT\ppt_system\ppt_system\PPT_PROJECTS_ENV_CONFIG.yml"
    parser = argparse.ArgumentParser(description="Generate blank slides_config template")
    parser.add_argument("--project", default="NCAER_MONTHLY_PPT", help="Project key from the YAML")
    parser.add_argument("--config", default=CONFIG_PATH, help="Path to the YAML config")
    args = parser.parse_args()
    
    paths = load_project_paths(args.config, args.project)
    template = paths["ppt_template"]
    slides_config = paths["slides_config"]
    # output = os.path.join(excel_dir, f"{args.project}_SLIDE_CONFIG_TEMPLATE.xlsx")
    
    # os.makedirs(excel_dir, exist_ok=True)
    print(f"Project : {args.project}")
    print(f"Template: {template}")
    print(f"Output  : {slides_config}")
    create_template(template, slides_config)