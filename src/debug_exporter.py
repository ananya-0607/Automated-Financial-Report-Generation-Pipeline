"""
debug_exporter.py — save chart data to Excel for inspection.

Usage in your main script
─────────────────────────
    from src.debug_exporter import export_chart_data

    slides = read_config(excel_path)
    # ... fetch data, populate cfg.x_labels / y1_values / y2_values ...

    proj = yaml_cfg["projects"]["NCAER_MONTHLY_PPT"]
    if proj.get("debug_chart_data"):
        export_chart_data(slides, proj["debug_chart_data_path"])

Output Excel structure
──────────────────────
  Sheet "00_Summary"  — one row per chart, all config fields + data stats
  Sheet "S01_C1" etc. — per-chart sheet: metadata block + actual x/y data
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── colour constants ──────────────────────────────────────────────────────────
_HDR_FILL  = "4472C4"   # blue  — summary header row
_HDR_FONT  = "FFFFFF"   # white — summary header text
_META_FILL = "E2EFDA"   # green — metadata key cells in chart sheets
_DATA_FILL = "D9E1F2"   # light blue — data column headers


def _bold(cell, color=None):
    cell.font = Font(bold=True, color=color or "000000")


def _header_style(cell):
    cell.font        = Font(bold=True, color=_HDR_FONT)
    cell.fill        = PatternFill("solid", fgColor=_HDR_FILL)
    cell.alignment   = Alignment(horizontal="center")


def _meta_style(cell):
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor=_META_FILL)


def _data_hdr_style(cell):
    cell.font = Font(bold=True)
    cell.fill = PatternFill("solid", fgColor=_DATA_FILL)


def _auto_width(ws, max_width=45):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(
            max_len + 2, max_width
        )


# ── main export function ──────────────────────────────────────────────────────

def export_chart_data(slides: list, output_path: str) -> None:
    """
    Write all chart data from a list of SlideConfig objects to an Excel file.

    Parameters
    ----------
    slides      : list[SlideConfig]  — returned by read_config(), with data
                  already populated in each ChartConfig
                  (cfg.x_labels, cfg.y1_values, cfg.y2_values)
    output_path : str — full path for the output .xlsx file
    """
    wb      = Workbook()
    ws_sum  = wb.active
    ws_sum.title = "00_Summary"

    # ── Summary sheet header ──────────────────────────────────────────────────
    SUMMARY_COLS = [
        "slide_number", "slide_title", "chart_position",
        "widget_id", "widget_id_older", "Fetching_Method",
        "y1_chart_type", "y1_label", "y1_divisor",
        "bar_color", "y_min", "y_max", "y_breaks",
        "y2_widget_id", "y2_chart_type", "y2_label",
        "y2_color", "y2_min", "y2_max",
        "x_grouping", "x_interval", "last_n_months",
        "start_year", "end_year", "growth_type", "aggregate",
        "show_values_all", "show_values_latest", "show_grid",
        "stack_widgets", "legend_loc", "legend_ncol", "legend_nrow",
        "source_override", "notes", "chart_title",
        "n_x_labels", "n_y1_points", "n_y2_points",
    ]
    ws_sum.append(SUMMARY_COLS)
    for cell in ws_sum[1]:
        _header_style(cell)

    total_charts = 0

    for slide in slides:
        for cfg in slide.charts:
            total_charts += 1

            x  = cfg.x_labels  or []
            y1 = cfg.y1_values or []
            y2 = cfg.y2_values or []

            # count data points (multi-series: length of first series)
            if y1 and isinstance(y1[0], list):
                n_y1 = len(y1[0])
            else:
                n_y1 = len(y1)

            # ── Summary row ───────────────────────────────────────────────
            ws_sum.append([
                slide.slide_number,
                slide.slide_title,
                cfg.chart_position,
                cfg.widget_id,
                cfg.widget_id_older,
                cfg.Fetching_Method,
                cfg.y1_chart_type,
                cfg.y1_label,
                cfg.y1_divisor,
                cfg.bar_color,
                cfg.y_min,
                cfg.y_max,
                cfg.y_breaks,
                cfg.y2_widget_id,
                cfg.y2_chart_type,
                cfg.y2_label,
                cfg.y2_color,
                cfg.y2_min,
                cfg.y2_max,
                cfg.x_grouping,
                cfg.x_interval,
                cfg.last_n_months,
                cfg.start_year,
                cfg.end_year,
                cfg.growth_type,
                cfg.aggregate,
                getattr(cfg, "show_values_all",    None),
                getattr(cfg, "show_values_latest", None),
                cfg.show_grid,
                cfg.stack_widgets,
                getattr(cfg, "legend_loc",  ""),
                getattr(cfg, "legend_ncol", 0),
                getattr(cfg, "legend_nrow", 0),
                cfg.source_override,
                cfg.notes,
                cfg.chart_title,
                len(x),
                n_y1,
                len(y2),
            ])

            # ── Per-chart data sheet ──────────────────────────────────────
            sheet_name = f"S{slide.slide_number:02d}_C{cfg.chart_position}"
            ws = wb.create_sheet(title=sheet_name)

            # metadata block
            meta_rows = [
                ("slide_number",   slide.slide_number),
                ("slide_title",    slide.slide_title),
                ("layout_name",    slide.layout_name),
                ("chart_position", cfg.chart_position),
                ("widget_id",      cfg.widget_id),
                ("y1_chart_type",  cfg.y1_chart_type),
                ("y1_label",       cfg.y1_label),
                ("y1_divisor",     cfg.y1_divisor),
                ("bar_color",      cfg.bar_color),
                ("y_min",          cfg.y_min),
                ("y_max",          cfg.y_max),
                ("y_breaks",       cfg.y_breaks),
                ("y2_widget_id",   cfg.y2_widget_id),
                ("y2_chart_type",  cfg.y2_chart_type),
                ("y2_label",       cfg.y2_label),
                ("y2_color",       cfg.y2_color),
                ("y2_min",         cfg.y2_min),
                ("y2_max",         cfg.y2_max),
                ("x_grouping",     cfg.x_grouping),
                ("x_interval",     cfg.x_interval),
                ("last_n_months",  cfg.last_n_months),
                ("growth_type",    cfg.growth_type),
                ("aggregate",      cfg.aggregate),
                ("show_values_all",    getattr(cfg, "show_values_all",    None)),
                ("show_values_latest", getattr(cfg, "show_values_latest", None)),
                ("show_grid",      cfg.show_grid),
                ("stack_widgets",  cfg.stack_widgets),
                ("chart_title",    cfg.chart_title),
                ("source",         cfg.source),
                ("n_x_labels",     len(x)),
                ("n_y1_points",    n_y1),
                ("n_y2_points",    len(y2)),
            ]
            for k, v in meta_rows:
                ws.append([k, str(v) if v is not None else ""])
                _meta_style(ws.cell(row=ws.max_row, column=1))

            ws.append([])   # blank separator row

            # ── data section ──────────────────────────────────────────────
            if y1 and isinstance(y1[0], list):
                # multi-series (stacked bar / multi-line)
                series_labels = (
                    getattr(cfg, "stack_labels", None) or
                    [f"Series_{i+1}" for i in range(len(y1))]
                )
                header = ["x_label"] + [str(l) for l in series_labels]
                ws.append(header)
                for cell in ws[ws.max_row]:
                    _data_hdr_style(cell)
                for j, xl in enumerate(x):
                    data_row = [xl] + [
                        (y1[s][j] if j < len(y1[s]) else "")
                        for s in range(len(y1))
                    ]
                    ws.append(data_row)

            else:
                # single-series or dual-axis
                y1_col = f"y1  [{cfg.y1_label or cfg.y1_chart_type}]"
                if cfg.y2_widget_id > 0:
                    y2_col = f"y2  [{cfg.y2_label or 'secondary'}]"
                    header = ["x_label", y1_col, y2_col]
                else:
                    header = ["x_label", y1_col]
                ws.append(header)
                for cell in ws[ws.max_row]:
                    _data_hdr_style(cell)
                for j, xl in enumerate(x):
                    v1 = y1[j] if j < len(y1) else ""
                    if cfg.y2_widget_id > 0:
                        v2 = y2[j] if j < len(y2) else ""
                        ws.append([xl, v1, v2])
                    else:
                        ws.append([xl, v1])

            # column widths
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 18
            if cfg.y2_widget_id > 0:
                ws.column_dimensions["C"].width = 18

    # auto-width summary sheet columns
    _auto_width(ws_sum)

    wb.save(output_path)
    print(f"\n[debug_exporter] Saved → {output_path}")
    print(f"  Slides: {len(slides)}   Charts: {total_charts}")
